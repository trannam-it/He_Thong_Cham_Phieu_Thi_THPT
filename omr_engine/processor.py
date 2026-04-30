# =============================================================================
# PROCESSOR - Ham chinh xu ly Phieu tra loi trac nghiem.
# ---------------------------------------------------------------------------
# NANG CAP:
#   - Xuat ca bo anh debug (cat, nhi phan, contour) cho tung phan.
#   - Luu anh goc vao Anh_chua_cham.
#   - Dung anh mau trong Anh_mau_phieu de refine warp (ECC alignment).
# =============================================================================

import cv2
import numpy as np
import json
import os
import shutil
from datetime import datetime

from .alignment import preprocess_and_align
from .omr_header import OMR_Header
from .omr_phan1 import OMR_Phan1
from .omr_phan2 import OMR_Phan2
from .omr_phan3 import OMR_Phan3
from .debug_exporter import export_debug_images


def imread_unicode(path):
    """Doc anh ho tro duong dan Unicode (Windows)."""
    try:
        data = np.fromfile(path, dtype=np.uint8)
        img = cv2.imdecode(data, cv2.IMREAD_COLOR)
        return img
    except Exception as e:
        print(f"Loi doc anh Unicode: {e}")
        return None


def _save_original_to_store(image_path, base_dir):
    """Copy ảnh gốc vào Anh_cham/Anh_chua_cham/ để lưu trữ.

    Nếu ảnh đã nằm trong thư mục này rồi thì không copy (tránh tự chép).
    """
    if not base_dir:
        return None
    dst_dir = os.path.join(base_dir, "Anh_cham", "Anh_chua_cham")
    os.makedirs(dst_dir, exist_ok=True)
    basename = os.path.basename(image_path)
    dst_path = os.path.join(dst_dir, basename)

    # Nếu nguồn và đích cùng 1 file thì không cần chép
    try:
        if os.path.abspath(image_path) == os.path.abspath(dst_path):
            return dst_path
    except Exception:
        pass

    try:
        shutil.copy2(image_path, dst_path)
        print(f"  [SAVE] Luu ban goc -> {dst_path}")
        return dst_path
    except Exception as e:
        print(f"  [SAVE] Khong luu duoc ban goc: {e}")
        return None


def process_full_omr(image_path, output_debug_dir=None, output_json_dir=None,
                      base_dir=None, save_original=True, export_debug_steps=True):
    """Cham 1 phieu tra loi trac nghiem.

    Args:
        image_path: duong dan anh phieu
        output_debug_dir: thu muc xuat anh debug (Anh_da_cham)
        output_json_dir : thu muc xuat JSON (Diem/JSON)
        base_dir: thu muc goc he thong (chua Anh_cham, Diem)
                  Dung de: (1) load anh mau de refine warp,
                           (2) luu ban goc vao Anh_chua_cham.
        save_original: True -> copy anh goc vao Anh_chua_cham
        export_debug_steps: True -> xuat tat ca anh debug trung gian

    Tra ve dict ket qua, hoac None neu khong xu ly duoc.
    """
    print("\n" + "#" * 70)
    print(f"# CHAM PHIEU: {os.path.basename(image_path)}")
    print("#" * 70)

    # Suy luận base_dir nếu không được cung cấp
    if base_dir is None and output_debug_dir:
        # output_debug_dir thường là base_dir/Anh_cham/Anh_da_cham
        # => base_dir = parent parent
        parent = os.path.dirname(os.path.dirname(output_debug_dir))
        if os.path.isdir(parent):
            base_dir = parent

    # ===== BUOC 0: LUU BAN GOC =====
    if save_original and base_dir:
        _save_original_to_store(image_path, base_dir)

    # ===== BUOC 1: DOC ANH =====
    img = imread_unicode(image_path)
    if img is None:
        print(f"Khong doc duoc anh: {image_path}")
        return None
    print(f"Da doc anh: shape={img.shape}")

    # ===== BUOC 2: TIEN XU LY + NAN CHINH =====
    print("\n-- BUOC 2: CHUAN HOA + NAN CHINH --")
    warped_color, warped_gray, info = preprocess_and_align(
        img, base_dir=base_dir, use_reference=True
    )
    print(f"  Warp method: {info['method']}")
    print(f"  Warped shape: {warped_color.shape}")
    if info.get('rotated', False):
        print(f"  Da sua nghieng: goc {info.get('skew_angle', 0):.2f} do "
              f"(nan perspective ve hinh chu nhat dung)")
    else:
        print(f"  Anh THANG DUNG -> khong xoay "
              f"(goc nghieng={info.get('skew_angle', 0):.2f} do)")

    # ===== BUOC 3: CHAY DETECTOR =====
    hdr = OMR_Header()
    result_hdr, debug_hdr = hdr.process(warped_gray)

    p1_det = OMR_Phan1()
    result_p1, cert_p1, debug_p1 = p1_det.process(warped_gray)

    p2_det = OMR_Phan2()
    result_p2, cert_p2, debug_p2 = p2_det.process(warped_gray)

    p3_det = OMR_Phan3()
    result_p3, cert_p3, debug_p3 = p3_det.process(warped_gray)

    # ===== BUOC 4: XUAT DEBUG + JSON =====
    print("\n-- BUOC 4: XUAT DEBUG + JSON --")
    canvas = warped_color.copy()
    hdr.draw_debug(canvas, result_hdr, debug_hdr)
    p1_det.draw_debug(canvas, result_p1, cert_p1, debug_p1)
    p2_det.draw_debug(canvas, result_p2, cert_p2, debug_p2)
    p3_det.draw_debug(canvas, result_p3, cert_p3, debug_p3)

    for pt in info['src_pts']:
        cv2.circle(canvas, (int(pt[0]), int(pt[1])), 6, (255, 0, 255), 2)
    cv2.putText(canvas, f"Align: {info['method']}",
                 (8, canvas.shape[0] - 12),
                 cv2.FONT_HERSHEY_SIMPLEX, 0.45, (128, 0, 128), 1)

    image_name = os.path.splitext(os.path.basename(image_path))[0]

    # Xuất debug images (legacy + multi-step)
    debug_paths = {}
    if output_debug_dir:
        os.makedirs(output_debug_dir, exist_ok=True)
        if export_debug_steps:
            debug_paths = export_debug_images(
                image_name=image_name,
                img_original=img,
                warped_color=warped_color,
                warped_gray=warped_gray,
                final_canvas=canvas,
                out_dir=output_debug_dir,
                info=info,
            )
            debug_path = debug_paths.get(
                'legacy_debug',
                os.path.join(output_debug_dir, f"{image_name}_debug.jpg")
            )
            print(f"  Anh debug (full set) -> {os.path.join(output_debug_dir, image_name + '_debug')}/")
        else:
            debug_path = os.path.join(output_debug_dir, f"{image_name}_debug.jpg")
            _ok = cv2.imencode('.jpg', canvas)
            if _ok[0]:
                _ok[1].tofile(debug_path)
    else:
        base = image_path.rsplit('.', 1)[0]
        debug_path = base + '_debug.jpg'
        cv2.imwrite(debug_path, canvas)
    print(f"  Anh debug chinh -> {debug_path}")

    result_full = {
        "image_name": os.path.basename(image_path),
        "header": {
            "sbd":            result_hdr['sbd'],
            "madt":           result_hdr['madt'],
            "sbd_digits":     result_hdr['sbd_digits'],
            "madt_digits":    result_hdr['madt_digits'],
            "sbd_certainty":  result_hdr['sbd_certainty'],
            "madt_certainty": result_hdr['madt_certainty'],
        },
        "phan1": {str(k): v for k, v in result_p1.items()},
        "phan2": {str(k): v for k, v in result_p2.items()},
        "phan3": {str(k): v for k, v in result_p3.items()},
        "align_method": info['method'],
        "skew_angle": float(info.get('skew_angle', 0)),
    }

    if output_json_dir:
        os.makedirs(output_json_dir, exist_ok=True)
        json_path = os.path.join(output_json_dir, f"{image_name}.json")
    else:
        base = image_path.rsplit('.', 1)[0]
        json_path = base + '_result.json'

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(result_full, f, ensure_ascii=False, indent=2)
    print(f"  Ket qua JSON -> {json_path}")

    # TOM TAT
    print("\n-- TOM TAT KET QUA --")
    print(f"  SBD      : {result_hdr['sbd']}")
    print(f"  Ma de    : {result_hdr['madt']}")
    print(f"  Phan I   : {sum(1 for v in result_p1.values() if v is not None)}/40 cau")
    p2_count = sum(1 for cau_val in result_p2.values()
                   for v in cau_val.values() if v is not None)
    print(f"  Phan II  : {p2_count}/32 y")
    p3_count = sum(1 for v in result_p3.values() if v)
    print(f"  Phan III : {p3_count}/6 cau")
    for cau in sorted(result_p3.keys()):
        print(f"     Cau {cau}: {result_p3[cau] or '(trong)'}")

    return {
        'image_name':       os.path.basename(image_path),
        'header':           result_hdr,
        'phan1':            result_p1,
        'phan2':            result_p2,
        'phan3':            result_p3,
        'debug_image':      debug_path,
        'debug_paths':      debug_paths,  # NEW: dict cac anh debug trung gian
        'json_result':      json_path,
        'result_json_data': result_full,
        'align_method':     info['method'],
        'skew_angle':       float(info.get('skew_angle', 0)),
    }


def process_batch(image_paths, base_dir):
    """Cham nhieu phieu mot luot, tra ve (all_results, xlsx_path)."""
    debug_dir = os.path.join(base_dir, "Anh_cham", "Anh_da_cham")
    json_dir  = os.path.join(base_dir, "Diem", "JSON")
    xlsx_dir  = os.path.join(base_dir, "Diem", "XLSX")
    os.makedirs(debug_dir, exist_ok=True)
    os.makedirs(json_dir, exist_ok=True)
    os.makedirs(xlsx_dir, exist_ok=True)

    all_results = []
    for img_path in image_paths:
        try:
            result = process_full_omr(
                img_path,
                output_debug_dir=debug_dir,
                output_json_dir=json_dir,
                base_dir=base_dir,
                save_original=True,
                export_debug_steps=True,
            )
            if result:
                all_results.append(result)
            else:
                all_results.append({
                    'image_name': os.path.basename(img_path),
                    'error': 'Khong xu ly duoc anh',
                })
        except Exception as e:
            import traceback
            traceback.print_exc()
            all_results.append({
                'image_name': os.path.basename(img_path),
                'error': str(e),
            })

    xlsx_path = _create_batch_excel(all_results, xlsx_dir)
    return all_results, xlsx_path


def _create_batch_excel(all_results, xlsx_dir):
    """Tao file Excel tong hop ket qua cham."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl chua cai, dang cai...")
        import subprocess
        subprocess.check_call(['pip', 'install', 'openpyxl'])
        import openpyxl

    from openpyxl.styles import Font, Alignment, PatternFill

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = os.path.join(xlsx_dir, f"KetQua_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ket qua cham phieu"

    headers = ["Ten anh", "SBD", "Ma de"]
    for i in range(1, 41):
        headers.append(f"P1_Cau{i}")
    for q in range(1, 9):
        for sub in ['a', 'b', 'c', 'd']:
            headers.append(f"DS_Cau{q}{sub}")
    for i in range(1, 7):
        headers.append(f"DienSo_Cau{i}")

    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for result in all_results:
        if 'error' in result:
            row = [result.get('image_name', ''), '', '']
            row.extend([''] * 40)
            row.extend([''] * 32)
            row.extend([''] * 6)
            ws.append(row)
            continue

        row = [
            result.get('image_name', ''),
            result.get('header', {}).get('sbd', ''),
            result.get('header', {}).get('madt', ''),
        ]
        p1 = result.get('phan1', {})
        for i in range(1, 41):
            val = p1.get(i, p1.get(str(i)))
            row.append(val if val else '')
        p2 = result.get('phan2', {})
        for q in range(1, 9):
            q_data = p2.get(q, p2.get(str(q), {}))
            for sub in ['a', 'b', 'c', 'd']:
                val = q_data.get(sub) if isinstance(q_data, dict) else None
                row.append(val if val else '')
        p3 = result.get('phan3', {})
        for i in range(1, 7):
            val = p3.get(i, p3.get(str(i)))
            row.append(val if val else '')

        ws.append(row)

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(20, max(8, max_len + 2))

    wb.save(xlsx_path)
    print(f"File Excel -> {xlsx_path}")
    return xlsx_path
